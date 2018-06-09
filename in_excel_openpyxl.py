# -*- coding: utf-8 -*-

from openpyxl import Workbook
from openpyxl.chart import (Reference,Series,BarChart3D,)
from openpyxl import load_workbook
import os
import sys
reload(sys)
sys.setdefaultencoding('utf-8')
#图表写入-------------------------------------指定2列
def in_excel(data_list,result_file,sheet_name,col_num,col1,col2,title):
    #打开文件
    if os.path.exists(result_file):
        wb = load_workbook(result_file)
    else:
        wb = Workbook()
    ws = wb.active
    #读sheet页
    sheet_list=wb.sheetnames #表里面有哪些sheet页
    print sheet_list
    # sheet = wb.get_sheet_by_name("Sheet3")
    # print(sheet["C"])    # (<Cell Sheet3.C1>, <Cell Sheet3.C2>, <Cell Sheet3.C3>, <Cell Sheet3.C4>, <Cell Sheet3.C5>, <Cell Sheet3.C6>, <Cell Sheet3.C7>, <Cell Sheet3.C8>, <Cell Sheet3.C9>, <Cell Sheet3.C10>)      <-第C列
    # print(sheet["4"])    # (<Cell Sheet3.A4>, <Cell Sheet3.B4>, <Cell Sheet3.C4>, <Cell Sheet3.D4>, <Cell Sheet3.E4>)     <-第4行
    # print(sheet["C4"].value)    # c4     <-第C4格的值
    # print(sheet.max_row)    # 10     <-最大行数
    # print(sheet.max_column)    # 5     <-最大列数
    #写入sheet页
    if sheet_name not in sheet_list:
        sheet_request=wb.create_sheet(sheet_name)
    else:
        sheet_request=wb[sheet_name]
    # rows = data_list
    # for row in rows:
    #     sheet_request.append(row)
    for i in range(len(data_list)):
        # print i[1],data_list[i][1]
        # print start_row,start_row + data_list[i][1]-1
        # sheet_request['A' + str(start_row)]=data_list[i][0]
        sheet_request[col1 + str(i+1)] = data_list[i][0]
        sheet_request[col2 + str(i+1)] = data_list[i][1]
    # 删除sheet
    # wb.remove('Sheet')
    # wb.remove_sheet('Sheet')
    #插入图表
    data = Reference(sheet_request, min_col=col_num, min_row=1, max_col=col_num, max_row=15)
    # titles = Reference(sheet_request, min_col=1, min_row=2, max_row=10)
    titles = Reference(sheet_request, min_col=title, min_row=2, max_row=15)
    chart = BarChart3D()
    chart.title = "用户行为分析"
    chart.add_data(data=data, titles_from_data=True)
    chart.set_categories(titles)
    sheet_request.add_chart(chart, "E5")
    #保存
    wb.save(result_file)
#合并单元格写入-------------------------------------第一列
def in_excel2(data_list,result_file,sheet_name):
    #打开文件
    if os.path.exists(result_file):
        wb = load_workbook(result_file)
    else:
        wb = Workbook()
    ws = wb.active
    #读sheet页
    sheet_list=wb.sheetnames #表里面有哪些sheet页
    if sheet_name not in sheet_list:
        sheet_request=wb.create_sheet(sheet_name)
    else:
        sheet_request=wb[sheet_name]
    # sheet_request.merge_cells(start_row=1,start_column=1,end_row=7,end_column=1)
    start_row=2
    # for i in data_list:
    for i in range(1,len(data_list)):
        # print i[1],data_list[i][1]
        print start_row,start_row + data_list[i][1]-1
        sheet_request['A' + str(start_row)]=data_list[i][0]
        sheet_request.merge_cells(start_row=int(start_row), end_row=int(start_row + data_list[i][1]-1), start_column=1, end_column=1)
        start_row = start_row + data_list[i][1]
    #保存
    wb.save(result_file)
#写入-------------------------------------
def in_excel3(data_list,result_file,sheet_name):
    #打开文件
    if os.path.exists(result_file):
        wb = load_workbook(result_file)
    else:
        wb = Workbook()
    ws = wb.active
    #读sheet页
    sheet_list=wb.sheetnames #表里面有哪些sheet页
    if sheet_name not in sheet_list:
        sheet_request=wb.create_sheet(sheet_name)
    else:
        sheet_request=wb[sheet_name]
    rows = data_list
    # for row in rows:
    #     print row
    sheet_request.append(data_list)
    #保存
    wb.save(result_file)
if __name__ == "__main__":
    data_list =[('utm', 'utm_num'), ('', 52182), ('baidu', 15695), ('360', 2339), ('baidu-zs', 499), ('aizhan', 117), ('5118', 5),
     ('mojing', 4), ('chinaz', 2), ('link', 1)]
    # data_list =[('referer', 'request', 'request2', 'ip'),("Apples", 1, 5),("Apples", 5, 2),("Apples", 4, 7),("Oranges", 6, 5)]

    result_file = "user.xlsx"
    sheet_tite = 'referer'
    in_excel(data_list, result_file, sheet_tite, 2, 'A', 'B', 1)